from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.shortcuts import get_object_or_404
from django.db.models import Q, Sum, Count, F
from django.db import models
from django.utils import timezone
from datetime import timedelta, date
from .models import (
    Category, Supplier, Product, StockTransaction, 
    ExpiryAlert, ProductTicket, Supermarket, Substore
)
from .serializers import (
    CategorySerializer, SupplierSerializer, ProductSerializer,
    ProductCreateSerializer, StockTransactionSerializer, ExpiryAlertSerializer,
    ProductTicketSerializer, BarcodeSearchSerializer, StockUpdateSerializer,
    DashboardStatsSerializer, SupermarketSerializer, SubstoreSerializer,
    SubstoreCreateSerializer
)
import json


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def halal_certified(self, request):
        """Get only Halal certified suppliers"""
        suppliers = self.queryset.filter(halal_certified=True)
        serializer = self.get_serializer(suppliers, many=True)
        return Response(serializer.data)


class ProductViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Show products belonging to the current supermarket and its substores
        try:
            supermarket = Supermarket.objects.get(user=self.request.user)
            # Include products from main supermarket and all its substores
            return Product.objects.filter(
                models.Q(supermarket=supermarket) | models.Q(substore__supermarket=supermarket),
                is_halal=True, 
                is_active=True
            )
        except Supermarket.DoesNotExist:
            # For users without supermarket profiles, show all products (admin view)
            if self.request.user.is_superuser:
                return Product.objects.filter(is_halal=True, is_active=True)
            return Product.objects.none()
    
    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'bulk_create':
            return ProductCreateSerializer
        return ProductSerializer
    
    def perform_create(self, serializer):
        # Automatically assign the product to the current user's supermarket
        try:
            supermarket = Supermarket.objects.get(user=self.request.user)
            serializer.save(supermarket=supermarket)
        except Supermarket.DoesNotExist:
            if self.request.user.is_superuser:
                serializer.save()  # Admin can create products without supermarket
            else:
                raise serializers.ValidationError("You must be registered as a supermarket to create products")
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create products in multiple stores (supermarket and/or substores)"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        product_data = request.data.get('product_data', {})
        target_stores = request.data.get('target_stores', [])  # List of store IDs
        add_to_all_stores = request.data.get('add_to_all_stores', False)
        
        if not product_data:
            return Response({'error': 'Product data is required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        created_products = []
        errors = []
        
        # Determine target stores
        stores_to_add = []
        
        if add_to_all_stores:
            # Add to main supermarket
            stores_to_add.append({'type': 'supermarket', 'id': supermarket.id})
            # Add to all substores
            for substore in supermarket.substores.filter(is_active=True):
                stores_to_add.append({'type': 'substore', 'id': substore.id})
        else:
            # Add to specific stores
            for store_info in target_stores:
                if store_info.get('type') == 'supermarket' and store_info.get('id') == supermarket.id:
                    stores_to_add.append(store_info)
                elif store_info.get('type') == 'substore':
                    try:
                        substore = Substore.objects.get(id=store_info.get('id'), supermarket=supermarket)
                        stores_to_add.append(store_info)
                    except Substore.DoesNotExist:
                        errors.append(f"Substore {store_info.get('id')} not found")
        
        # Create products for each target store
        for store_info in stores_to_add:
            try:
                product_data_copy = product_data.copy()
                
                if store_info['type'] == 'supermarket':
                    product_data_copy['supermarket'] = supermarket.id
                    product_data_copy['substore'] = None
                else:  # substore
                    product_data_copy['supermarket'] = None
                    product_data_copy['substore'] = store_info['id']
                
                serializer = ProductCreateSerializer(data=product_data_copy, context={'request': request})
                if serializer.is_valid():
                    product = serializer.save()
                    created_products.append({
                        'product': ProductSerializer(product).data,
                        'store_type': store_info['type'],
                        'store_id': store_info['id']
                    })
                else:
                    errors.append(f"Store {store_info['type']}:{store_info['id']} - {serializer.errors}")
            except Exception as e:
                errors.append(f"Store {store_info['type']}:{store_info['id']} - {str(e)}")
        
        return Response({
            'created_products': created_products,
            'total_created': len(created_products),
            'errors': errors,
            'success': len(created_products) > 0
        }, status=status.HTTP_201_CREATED if created_products else status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def scan_barcode(self, request):
        """Scan barcode or QR code to find product"""
        serializer = BarcodeSearchSerializer(data=request.data)
        if serializer.is_valid():
            code = serializer.validated_data['code']
            scan_type = serializer.validated_data['scan_type']
            
            try:
                if scan_type == 'BARCODE':
                    product = Product.objects.get(barcode=code, is_halal=True, is_active=True)
                else:  # QR_CODE
                    # Parse QR code data (assuming it contains product ID or barcode)
                    try:
                        qr_data = json.loads(code.replace("'", '"'))
                        product_id = qr_data.get('id')
                        barcode = qr_data.get('barcode')
                        
                        if product_id:
                            product = Product.objects.get(id=product_id, is_halal=True, is_active=True)
                        elif barcode:
                            product = Product.objects.get(barcode=barcode, is_halal=True, is_active=True)
                        else:
                            return Response({'error': 'Invalid QR code format'}, 
                                          status=status.HTTP_400_BAD_REQUEST)
                    except (json.JSONDecodeError, ValueError):
                        # Try as direct barcode search
                        product = Product.objects.get(barcode=code, is_halal=True, is_active=True)
                
                product_serializer = ProductSerializer(product)
                return Response({
                    'product': product_serializer.data,
                    'scan_successful': True,
                    'message': f'Halal product found: {product.name}'
                })
                
            except Product.DoesNotExist:
                return Response({
                    'error': 'Product not found or not Halal certified',
                    'scan_successful': False,
                    'message': 'This product is not in our Halal inventory'
                }, status=status.HTTP_404_NOT_FOUND)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Get products expiring within 30 days"""
        thirty_days_from_now = timezone.now().date() + timedelta(days=30)
        products = self.queryset.filter(
            expiry_date__lte=thirty_days_from_now,
            expiry_date__gt=timezone.now().date()
        )
        serializer = self.get_serializer(products, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def expired(self, request):
        """Get expired products"""
        products = self.queryset.filter(expiry_date__lt=timezone.now().date())
        serializer = self.get_serializer(products, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Get products with low stock"""
        products = self.queryset.filter(current_stock__lte=F('minimum_stock'))
        serializer = self.get_serializer(products, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def update_stock(self, request, pk=None):
        """Update product stock"""
        product = get_object_or_404(Product, pk=pk, is_halal=True, is_active=True)
        serializer = StockUpdateSerializer(data=request.data)
        
        if serializer.is_valid():
            transaction_type = serializer.validated_data['transaction_type']
            quantity = serializer.validated_data['quantity']
            reason = serializer.validated_data.get('reason', '')
            
            previous_stock = product.current_stock
            
            if transaction_type == 'IN':
                product.current_stock += quantity
            elif transaction_type == 'OUT':
                if product.current_stock < quantity:
                    return Response({
                        'error': 'Insufficient stock',
                        'available_stock': product.current_stock
                    }, status=status.HTTP_400_BAD_REQUEST)
                product.current_stock -= quantity
            elif transaction_type == 'ADJUSTMENT':
                product.current_stock = quantity
            
            product.save()
            
            # Create stock transaction record
            StockTransaction.objects.create(
                product=product,
                transaction_type=transaction_type,
                quantity=quantity,
                previous_stock=previous_stock,
                new_stock=product.current_stock,
                reason=reason,
                user=request.user
            )
            
            product_serializer = ProductSerializer(product)
            return Response({
                'product': product_serializer.data,
                'message': 'Stock updated successfully'
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def generate_ticket(self, request, pk=None):
        """Generate product ticket/label"""
        product = get_object_or_404(Product, pk=pk, is_halal=True, is_active=True)
        
        ticket_data = {
            'product_name': product.name,
            'sku': product.sku,
            'barcode': product.barcode,
            'barcode_image': product.barcode_image,
            'qr_code_image': product.qr_code_image,
            'price': str(product.price),
            'expiry_date': product.expiry_date.strftime('%Y-%m-%d') if product.expiry_date else None,
            'halal_status': 'HALAL CERTIFIED',
            'category': product.category.name,
            'supplier': product.supplier.name,
            'generated_at': timezone.now().isoformat()
        }
        
        ticket = ProductTicket.objects.create(
            product=product,
            ticket_data=ticket_data,
            created_by=request.user
        )
        
        serializer = ProductTicketSerializer(ticket)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import products from Excel file"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Validate file type
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            import openpyxl
            from openpyxl import load_workbook
            
            # Get current supermarket
            try:
                supermarket = Supermarket.objects.get(user=request.user)
            except Supermarket.DoesNotExist:
                return Response({'error': 'You must be registered as a supermarket'}, 
                              status=status.HTTP_400_BAD_REQUEST)
            
            # Load workbook
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            imported_products = []
            errors = []
            
            # Expected columns: Name, SKU, Category, Supplier, Price, Current Stock, Minimum Stock, Expiry Date, Description
            headers = [cell.value for cell in worksheet[1]]
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # Skip empty rows
                        continue
                    
                    # Map row data to product fields
                    product_data = {}
                    for i, header in enumerate(headers):
                        if i < len(row) and header:
                            header_lower = header.lower().strip()
                            if header_lower in ['name', 'product name']:
                                product_data['name'] = row[i]
                            elif header_lower in ['sku', 'product code']:
                                product_data['sku'] = row[i]
                            elif header_lower in ['category']:
                                product_data['category_name'] = row[i]
                            elif header_lower in ['supplier']:
                                product_data['supplier_name'] = row[i]
                            elif header_lower in ['price']:
                                product_data['price'] = float(row[i]) if row[i] else 0
                            elif header_lower in ['current stock', 'stock']:
                                product_data['current_stock'] = int(row[i]) if row[i] else 0
                            elif header_lower in ['minimum stock', 'min stock']:
                                product_data['minimum_stock'] = int(row[i]) if row[i] else 0
                            elif header_lower in ['expiry date', 'expiry']:
                                if row[i]:
                                    from datetime import datetime
                                    if isinstance(row[i], datetime):
                                        product_data['expiry_date'] = row[i].date()
                                    else:
                                        product_data['expiry_date'] = datetime.strptime(str(row[i]), '%Y-%m-%d').date()
                            elif header_lower in ['description']:
                                product_data['description'] = row[i]
                    
                    # Validate required fields
                    if not product_data.get('name'):
                        errors.append(f"Row {row_num}: Product name is required")
                        continue
                    
                    # Get or create category
                    category = None
                    if product_data.get('category_name'):
                        category, created = Category.objects.get_or_create(
                            name=product_data['category_name'],
                            defaults={'description': f'Auto-created from Excel import'}
                        )
                    
                    # Get or create supplier
                    supplier = None
                    if product_data.get('supplier_name'):
                        supplier, created = Supplier.objects.get_or_create(
                            name=product_data['supplier_name'],
                            defaults={
                                'contact_person': 'Unknown',
                                'email': 'unknown@example.com',
                                'phone': 'Unknown',
                                'address': 'Unknown',
                                'halal_certified': True  # Assume halal for import
                            }
                        )
                    
                    # Create product
                    product = Product.objects.create(
                        name=product_data['name'],
                        sku=product_data.get('sku', f'SKU{row_num}'),
                        category=category,
                        supplier=supplier,
                        price=product_data.get('price', 0),
                        current_stock=product_data.get('current_stock', 0),
                        minimum_stock=product_data.get('minimum_stock', 0),
                        expiry_date=product_data.get('expiry_date'),
                        description=product_data.get('description', ''),
                        supermarket=supermarket,
                        is_halal=True,
                        is_active=True
                    )
                    
                    imported_products.append(ProductSerializer(product).data)
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
            return Response({
                'message': f'Successfully imported {len(imported_products)} products',
                'imported_products': imported_products,
                'total_imported': len(imported_products),
                'errors': errors,
                'success': len(imported_products) > 0
            }, status=status.HTTP_201_CREATED if imported_products else status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            return Response({'error': f'Failed to process Excel file: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class StockTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = StockTransaction.objects.all()
    serializer_class = StockTransactionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        product_id = self.request.query_params.get('product_id', None)
        if product_id:
            queryset = queryset.filter(product__id=product_id)
        return queryset.order_by('-created_at')


class ExpiryAlertViewSet(viewsets.ModelViewSet):
    queryset = ExpiryAlert.objects.all()
    serializer_class = ExpiryAlertSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return super().get_queryset().filter(is_read=False).order_by('-created_at')
    
    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """Mark alert as read"""
        alert = get_object_or_404(ExpiryAlert, pk=pk)
        alert.is_read = True
        alert.save()
        serializer = self.get_serializer(alert)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        """Mark all alerts as read"""
        updated = ExpiryAlert.objects.filter(is_read=False).update(is_read=True)
        return Response({'message': f'{updated} alerts marked as read'})
    
    @action(detail=False, methods=['post'])
    def generate_alerts(self, request):
        """Manually generate expiry alerts"""
        thirty_days_from_now = timezone.now().date() + timedelta(days=30)
        today = timezone.now().date()
        
        # Products expiring soon
        expiring_products = Product.objects.filter(
            expiry_date__lte=thirty_days_from_now,
            expiry_date__gt=today,
            is_halal=True,
            is_active=True
        )
        
        # Expired products
        expired_products = Product.objects.filter(
            expiry_date__lt=today,
            is_halal=True,
            is_active=True
        )
        
        alerts_created = 0
        
        # Create alerts for expiring products
        for product in expiring_products:
            alert, created = ExpiryAlert.objects.get_or_create(
                product=product,
                alert_type='EXPIRING_SOON'
            )
            if created:
                alerts_created += 1
        
        # Create alerts for expired products
        for product in expired_products:
            alert, created = ExpiryAlert.objects.get_or_create(
                product=product,
                alert_type='EXPIRED'
            )
            if created:
                alerts_created += 1
        
        return Response({
            'message': f'{alerts_created} new alerts created',
            'expiring_soon_count': expiring_products.count(),
            'expired_count': expired_products.count()
        })


class ProductTicketViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ProductTicket.objects.all()
    serializer_class = ProductTicketSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        product_id = self.request.query_params.get('product_id', None)
        if product_id:
            queryset = queryset.filter(product__id=product_id)
        return queryset.order_by('-created_at')


class SupermarketViewSet(viewsets.ModelViewSet):
    queryset = Supermarket.objects.all()
    serializer_class = SupermarketSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Regular users can only see their own supermarket
        # Superusers can see all supermarkets
        if self.request.user.is_superuser:
            return Supermarket.objects.all()
        
        try:
            supermarket = Supermarket.objects.get(user=self.request.user)
            return Supermarket.objects.filter(id=supermarket.id)
        except Supermarket.DoesNotExist:
            return Supermarket.objects.none()
    
    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Verify a supermarket (admin only)"""
        if not request.user.is_superuser:
            return Response({'error': 'Only admins can verify supermarkets'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        supermarket = get_object_or_404(Supermarket, pk=pk)
        supermarket.is_verified = True
        supermarket.verified_by = request.user
        supermarket.verified_at = timezone.now()
        supermarket.save()
        
        serializer = self.get_serializer(supermarket)
        return Response({
            'message': f'Supermarket {supermarket.name} has been verified',
            'supermarket': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def verified(self, request):
        """Get only verified supermarkets"""
        queryset = self.get_queryset().filter(is_verified=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class DashboardViewSet(viewsets.ViewSet):
    """Dashboard statistics and overview"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get dashboard statistics for supermarket and all substores"""
        today = timezone.now().date()
        thirty_days_from_now = today + timedelta(days=30)
        
        try:
            supermarket = Supermarket.objects.get(user=request.user)
            # Get products from supermarket and all its substores
            products = Product.objects.filter(
                models.Q(supermarket=supermarket) | models.Q(substore__supermarket=supermarket),
                is_halal=True, 
                is_active=True
            )
        except Supermarket.DoesNotExist:
            if request.user.is_superuser:
                products = Product.objects.filter(is_halal=True, is_active=True)
            else:
                products = Product.objects.none()
        
        # Basic counts
        total_products = products.count()
        total_halal_products = total_products  # Since we only show Halal products
        
        # Stock-related stats
        low_stock_products = products.filter(current_stock__lte=F('minimum_stock')).count()
        
        # Expiry-related stats
        expiring_soon_products = products.filter(
            expiry_date__lte=thirty_days_from_now,
            expiry_date__gt=today
        ).count()
        expired_products = products.filter(expiry_date__lt=today).count()
        
        # Stock value
        total_stock_value = products.aggregate(
            total=Sum(F('current_stock') * F('cost_price'))
        )['total'] or 0
        
        # Categories and suppliers
        categories_count = Category.objects.count()
        suppliers_count = Supplier.objects.filter(halal_certified=True).count()
        
        # Stock status breakdown
        stock_stats = {
            'out_of_stock_count': products.filter(current_stock=0).count(),
            'low_stock_count': products.filter(
                current_stock__lte=F('minimum_stock'),
                current_stock__gt=0
            ).count(),
            'normal_stock_count': products.filter(
                current_stock__gt=F('minimum_stock'),
                current_stock__lt=F('maximum_stock')
            ).count(),
            'overstock_count': products.filter(
                current_stock__gte=F('maximum_stock')
            ).count(),
        }
        
        # Recent activities (filtered by supermarket)
        recent_transactions = StockTransaction.objects.select_related('product', 'user').filter(
            models.Q(product__supermarket=supermarket) | models.Q(product__substore__supermarket=supermarket)
        )[:10] if 'supermarket' in locals() else StockTransaction.objects.select_related('product', 'user')[:10]
        
        recent_alerts = ExpiryAlert.objects.select_related('product').filter(
            models.Q(product__supermarket=supermarket) | models.Q(product__substore__supermarket=supermarket),
            is_read=False
        )[:10] if 'supermarket' in locals() else ExpiryAlert.objects.select_related('product').filter(is_read=False)[:10]
        
        # Store-wise breakdown
        store_breakdown = []
        if 'supermarket' in locals():
            # Main supermarket stats
            main_products = products.filter(supermarket=supermarket, substore__isnull=True)
            store_breakdown.append({
                'store_type': 'supermarket',
                'store_id': supermarket.id,
                'store_name': supermarket.name,
                'total_products': main_products.count(),
                'total_stock_value': main_products.aggregate(
                    total=Sum(F('current_stock') * F('cost_price'))
                )['total'] or 0,
                'low_stock_count': main_products.filter(current_stock__lte=F('minimum_stock')).count(),
                'expiring_soon_count': main_products.filter(
                    expiry_date__lte=thirty_days_from_now,
                    expiry_date__gt=today
                ).count(),
            })
            
            # Substores stats
            for substore in supermarket.substores.filter(is_active=True):
                substore_products = products.filter(substore=substore)
                store_breakdown.append({
                    'store_type': 'substore',
                    'store_id': substore.id,
                    'store_name': substore.name,
                    'total_products': substore_products.count(),
                    'total_stock_value': substore_products.aggregate(
                        total=Sum(F('current_stock') * F('cost_price'))
                    )['total'] or 0,
                    'low_stock_count': substore_products.filter(current_stock__lte=F('minimum_stock')).count(),
                    'expiring_soon_count': substore_products.filter(
                        expiry_date__lte=thirty_days_from_now,
                        expiry_date__gt=today
                    ).count(),
                })
        
        stats_data = {
            'total_products': total_products,
            'total_halal_products': total_halal_products,
            'low_stock_products': low_stock_products,
            'expiring_soon_products': expiring_soon_products,
            'expired_products': expired_products,
            'total_stock_value': total_stock_value,
            'categories_count': categories_count,
            'suppliers_count': suppliers_count,
            'store_breakdown': store_breakdown,
            **stock_stats,
            'recent_stock_transactions': StockTransactionSerializer(recent_transactions, many=True).data,
            'recent_expiry_alerts': ExpiryAlertSerializer(recent_alerts, many=True).data
        }
        
        return Response(stats_data)
    
    @action(detail=False, methods=['get'])
    def store_specific_stats(self, request):
        """Get statistics for a specific store (supermarket or substore)"""
        store_type = request.query_params.get('store_type')  # 'supermarket' or 'substore'
        store_id = request.query_params.get('store_id')
        
        if not store_type or not store_id:
            return Response({'error': 'store_type and store_id are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            if not request.user.is_superuser:
                return Response({'error': 'You must be registered as a supermarket'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        
        today = timezone.now().date()
        thirty_days_from_now = today + timedelta(days=30)
        
        # Get products for specific store
        if store_type == 'supermarket':
            if 'supermarket' in locals() and str(supermarket.id) != store_id:
                return Response({'error': 'Access denied to this supermarket'}, 
                              status=status.HTTP_403_FORBIDDEN)
            products = Product.objects.filter(supermarket_id=store_id, substore__isnull=True, is_active=True)
            store_name = supermarket.name if 'supermarket' in locals() else 'Supermarket'
        else:  # substore
            try:
                if 'supermarket' in locals():
                    substore = Substore.objects.get(id=store_id, supermarket=supermarket)
                else:
                    substore = Substore.objects.get(id=store_id)
                products = Product.objects.filter(substore=substore, is_active=True)
                store_name = substore.name
            except Substore.DoesNotExist:
                return Response({'error': 'Substore not found'}, 
                              status=status.HTTP_404_NOT_FOUND)
        
        # Calculate statistics
        total_products = products.count()
        low_stock_products = products.filter(current_stock__lte=F('minimum_stock')).count()
        expiring_soon_products = products.filter(
            expiry_date__lte=thirty_days_from_now,
            expiry_date__gt=today
        ).count()
        expired_products = products.filter(expiry_date__lt=today).count()
        total_stock_value = products.aggregate(
            total=Sum(F('current_stock') * F('cost_price'))
        )['total'] or 0
        
        # Stock status breakdown
        stock_stats = {
            'out_of_stock_count': products.filter(current_stock=0).count(),
            'low_stock_count': products.filter(
                current_stock__lte=F('minimum_stock'),
                current_stock__gt=0
            ).count(),
            'normal_stock_count': products.filter(
                current_stock__gt=F('minimum_stock'),
                current_stock__lt=F('maximum_stock')
            ).count(),
            'overstock_count': products.filter(
                current_stock__gte=F('maximum_stock')
            ).count(),
        }
        
        # Recent activities for this store
        recent_transactions = StockTransaction.objects.select_related('product', 'user').filter(
            product__in=products
        )[:10]
        
        recent_alerts = ExpiryAlert.objects.select_related('product').filter(
            product__in=products,
            is_read=False
        )[:10]
        
        return Response({
            'store_type': store_type,
            'store_id': store_id,
            'store_name': store_name,
            'total_products': total_products,
            'low_stock_products': low_stock_products,
            'expiring_soon_products': expiring_soon_products,
            'expired_products': expired_products,
            'total_stock_value': total_stock_value,
            **stock_stats,
            'recent_stock_transactions': StockTransactionSerializer(recent_transactions, many=True).data,
            'recent_expiry_alerts': ExpiryAlertSerializer(recent_alerts, many=True).data
        })
    
    @action(detail=False, methods=['get'])
    def alerts_summary(self, request):
        """Get alerts summary for dashboard"""
        today = timezone.now().date()
        
        # Get unread alerts grouped by type
        expiring_alerts = ExpiryAlert.objects.filter(
            alert_type='EXPIRING_SOON',
            is_read=False
        ).select_related('product')
        
        expired_alerts = ExpiryAlert.objects.filter(
            alert_type='EXPIRED',
            is_read=False
        ).select_related('product')
        
        return Response({
            'expiring_soon': {
                'count': expiring_alerts.count(),
                'alerts': ExpiryAlertSerializer(expiring_alerts[:5], many=True).data
            },
            'expired': {
                'count': expired_alerts.count(),
                'alerts': ExpiryAlertSerializer(expired_alerts[:5], many=True).data
            },
            'total_unread': expiring_alerts.count() + expired_alerts.count()
        })


class SupermarketViewSet(viewsets.ModelViewSet):
    """ViewSet for managing supermarkets"""
    permission_classes = [IsAuthenticated]
    serializer_class = SupermarketSerializer
    
    def get_queryset(self):
        # Only show the current user's supermarket
        try:
            supermarket = Supermarket.objects.get(user=self.request.user)
            return Supermarket.objects.filter(id=supermarket.id)
        except Supermarket.DoesNotExist:
            if self.request.user.is_superuser:
                return Supermarket.objects.all()
            return Supermarket.objects.none()
    
    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def register(self, request):
        """Register a new supermarket - same as auth_views.register_supermarket"""
        from .auth_views import register_supermarket
        return register_supermarket(request)
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        """Get current user's supermarket details"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
            return Response(SupermarketSerializer(supermarket).data)
        except Supermarket.DoesNotExist:
            return Response({'error': 'Supermarket profile not found'}, 
                          status=status.HTTP_404_NOT_FOUND)


class SubstoreViewSet(viewsets.ModelViewSet):
    """ViewSet for managing substores"""
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Only show substores belonging to the current user's supermarket
        try:
            supermarket = Supermarket.objects.get(user=self.request.user)
            return Substore.objects.filter(supermarket=supermarket)
        except Supermarket.DoesNotExist:
            if self.request.user.is_superuser:
                return Substore.objects.all()
            return Substore.objects.none()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return SubstoreCreateSerializer
        return SubstoreSerializer
    
    def perform_create(self, serializer):
        # Automatically assign the substore to the current user's supermarket
        try:
            supermarket = Supermarket.objects.get(user=self.request.user)
            serializer.save(supermarket=supermarket)
        except Supermarket.DoesNotExist:
            if self.request.user.is_superuser:
                # Admin can create substores for any supermarket
                serializer.save()
            else:
                raise serializers.ValidationError("You must be registered as a supermarket to create substores")
    
    @action(detail=True, methods=['get'])
    def products(self, request, pk=None):
        """Get all products for a specific substore"""
        substore = get_object_or_404(Substore, pk=pk)
        products = Product.objects.filter(substore=substore, is_active=True)
        serializer = ProductSerializer(products, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """Get statistics for a specific substore"""
        substore = get_object_or_404(Substore, pk=pk)
        products = Product.objects.filter(substore=substore, is_active=True)
        
        stats = {
            'total_products': products.count(),
            'total_stock_value': substore.total_stock_value,
            'low_stock_products': products.filter(current_stock__lte=F('minimum_stock')).count(),
            'expiring_soon_products': products.filter(
                expiry_date__lte=timezone.now().date() + timedelta(days=30),
                expiry_date__gt=timezone.now().date()
            ).count(),
            'expired_products': products.filter(expiry_date__lt=timezone.now().date()).count(),
        }
        
        return Response(stats)


class ExcelImportViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Excel imports"""
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Only show imports belonging to the current user's supermarket
        try:
            supermarket = Supermarket.objects.get(user=self.request.user)
            return ExcelImport.objects.filter(
                models.Q(supermarket=supermarket) | models.Q(substore__supermarket=supermarket)
            )
        except Supermarket.DoesNotExist:
            if self.request.user.is_superuser:
                return ExcelImport.objects.all()
            return ExcelImport.objects.filter(uploaded_by=self.request.user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ExcelImportCreateSerializer
        return ExcelImportSerializer
    
    def create(self, request, *args, **kwargs):
        """Create a new Excel import and process it"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Get user's supermarket
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Determine target store
        target_store = serializer.validated_data['target_store']
        substore = None
        
        if target_store == 'substore':
            substore_id = serializer.validated_data['substore_id']
            try:
                substore = Substore.objects.get(id=substore_id, supermarket=supermarket)
            except Substore.DoesNotExist:
                return Response({'error': 'Substore not found'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        
        # Create Excel import record
        excel_import = ExcelImport.objects.create(
            file_name=serializer.validated_data['file_name'],
            file_data=serializer.validated_data['file_data'],
            uploaded_by=request.user,
            supermarket=supermarket if target_store == 'supermarket' else None,
            substore=substore
        )
        
        # Process the Excel file asynchronously (in a real app, use Celery)
        try:
            ExcelImport.process_excel_file(excel_import)
        except Exception as e:
            excel_import.status = 'FAILED'
            excel_import.error_log = str(e)
            excel_import.save()
        
        return Response(ExcelImportSerializer(excel_import).data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['get'])
    def status(self, request, pk=None):
        """Get import status"""
        excel_import = get_object_or_404(ExcelImport, pk=pk)
        return Response({
            'status': excel_import.status,
            'total_rows': excel_import.total_rows,
            'processed_rows': excel_import.processed_rows,
            'successful_imports': excel_import.successful_imports,
            'failed_imports': excel_import.failed_imports,
            'error_log': excel_import.error_log
        })


class ImageImportViewSet(viewsets.ModelViewSet):
    """ViewSet for managing Image imports"""
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Only show imports belonging to the current user's supermarket
        try:
            supermarket = Supermarket.objects.get(user=self.request.user)
            return ImageImport.objects.filter(
                models.Q(supermarket=supermarket) | models.Q(substore__supermarket=supermarket)
            )
        except Supermarket.DoesNotExist:
            if self.request.user.is_superuser:
                return ImageImport.objects.all()
            return ImageImport.objects.filter(uploaded_by=self.request.user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ImageImportCreateSerializer
        return ImageImportSerializer
    
    def create(self, request, *args, **kwargs):
        """Create a new Image import and process it"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Get user's supermarket
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Determine target store
        target_store = serializer.validated_data['target_store']
        substore = None
        
        if target_store == 'substore':
            substore_id = serializer.validated_data['substore_id']
            try:
                substore = Substore.objects.get(id=substore_id, supermarket=supermarket)
            except Substore.DoesNotExist:
                return Response({'error': 'Substore not found'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        
        # Create Image import record
        image_import = ImageImport.objects.create(
            image_name=serializer.validated_data['image_name'],
            image_data=serializer.validated_data['image_data'],
            uploaded_by=request.user,
            supermarket=supermarket if target_store == 'supermarket' else None,
            substore=substore
        )
        
        # Process the image file
        try:
            extracted_data = ImageImport.process_image_file(image_import)
            return Response({
                'import_data': ImageImportSerializer(image_import).data,
                'extracted_data': extracted_data
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            image_import.status = 'FAILED'
            image_import.error_log = str(e)
            image_import.save()
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def create_product_from_extraction(self, request, pk=None):
        """Create a product from extracted image data"""
        image_import = get_object_or_404(ImageImport, pk=pk)
        
        if image_import.status != 'COMPLETED':
            return Response({'error': 'Image processing not completed'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Get product data from request
        product_data = request.data
        
        # Add import information
        product_data.update({
            'imported_from_image': True,
            'import_batch_id': str(image_import.batch_id),
            'supermarket': image_import.supermarket.id if image_import.supermarket else None,
            'substore': image_import.substore.id if image_import.substore else None,
        })
        
        # Create product
        serializer = ProductCreateSerializer(data=product_data, context={'request': request})
        if serializer.is_valid():
            product = serializer.save()
            return Response(ProductSerializer(product).data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class POSIntegrationViewSet(viewsets.ViewSet):
    """POS Integration endpoints for external systems"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def products_sync(self, request):
        """Get all products for POS synchronization"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
            products = Product.objects.filter(
                models.Q(supermarket=supermarket) | models.Q(substore__supermarket=supermarket),
                is_halal=True, 
                is_active=True
            )
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Format products for POS system
        pos_products = []
        for product in products:
            pos_products.append({
                'id': str(product.id),
                'sku': product.sku,
                'name': product.name,
                'barcode': product.barcode,
                'price': float(product.price),
                'cost_price': float(product.cost_price),
                'current_stock': product.current_stock,
                'category': product.category.name,
                'supplier': product.supplier.name,
                'is_halal': product.is_halal,
                'store_location': product.store_location,
                'store_type': 'substore' if product.substore else 'supermarket',
                'store_id': product.substore.id if product.substore else product.supermarket.id,
                'last_updated': product.updated_at.isoformat(),
            })
        
        return Response({
            'products': pos_products,
            'total_count': len(pos_products),
            'sync_timestamp': timezone.now().isoformat()
        })
    
    @action(detail=False, methods=['post'])
    def stock_update_from_pos(self, request):
        """Update stock levels from POS system"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        updates = request.data.get('updates', [])
        if not updates:
            return Response({'error': 'No updates provided'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        successful_updates = []
        failed_updates = []
        
        for update in updates:
            try:
                product_id = update.get('product_id')
                new_stock = update.get('new_stock')
                transaction_type = update.get('transaction_type', 'ADJUSTMENT')
                reason = update.get('reason', 'POS System Update')
                
                # Find product
                product = Product.objects.get(
                    Q(id=product_id) & 
                    (Q(supermarket=supermarket) | Q(substore__supermarket=supermarket)) &
                    Q(is_active=True)
                )
                
                previous_stock = product.current_stock
                product.current_stock = new_stock
                product.save()
                
                # Create stock transaction record
                StockTransaction.objects.create(
                    product=product,
                    transaction_type=transaction_type,
                    quantity=abs(new_stock - previous_stock),
                    previous_stock=previous_stock,
                    new_stock=new_stock,
                    reason=reason,
                    user=request.user
                )
                
                successful_updates.append({
                    'product_id': product_id,
                    'previous_stock': previous_stock,
                    'new_stock': new_stock,
                    'status': 'success'
                })
                
            except Product.DoesNotExist:
                failed_updates.append({
                    'product_id': update.get('product_id'),
                    'error': 'Product not found',
                    'status': 'failed'
                })
            except Exception as e:
                failed_updates.append({
                    'product_id': update.get('product_id'),
                    'error': str(e),
                    'status': 'failed'
                })
        
        return Response({
            'successful_updates': successful_updates,
            'failed_updates': failed_updates,
            'total_processed': len(updates),
            'success_count': len(successful_updates),
            'failure_count': len(failed_updates)
        })
    
    @action(detail=False, methods=['post'])
    def sales_data_from_pos(self, request):
        """Receive sales data from POS system"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        sales_data = request.data.get('sales', [])
        if not sales_data:
            return Response({'error': 'No sales data provided'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        processed_sales = []
        failed_sales = []
        
        for sale in sales_data:
            try:
                product_id = sale.get('product_id')
                quantity_sold = sale.get('quantity_sold', 0)
                sale_price = sale.get('sale_price', 0)
                sale_timestamp = sale.get('timestamp')
                
                # Find product and update stock
                product = Product.objects.get(
                    Q(id=product_id) & 
                    (Q(supermarket=supermarket) | Q(substore__supermarket=supermarket)) &
                    Q(is_active=True)
                )
                
                if product.current_stock >= quantity_sold:
                    previous_stock = product.current_stock
                    product.current_stock -= quantity_sold
                    product.save()
                    
                    # Create stock transaction record
                    StockTransaction.objects.create(
                        product=product,
                        transaction_type='OUT',
                        quantity=quantity_sold,
                        previous_stock=previous_stock,
                        new_stock=product.current_stock,
                        reason=f'POS Sale - ${sale_price}',
                        user=request.user
                    )
                    
                    processed_sales.append({
                        'product_id': product_id,
                        'quantity_sold': quantity_sold,
                        'remaining_stock': product.current_stock,
                        'status': 'success'
                    })
                else:
                    failed_sales.append({
                        'product_id': product_id,
                        'error': f'Insufficient stock. Available: {product.current_stock}, Requested: {quantity_sold}',
                        'status': 'failed'
                    })
                    
            except Product.DoesNotExist:
                failed_sales.append({
                    'product_id': sale.get('product_id'),
                    'error': 'Product not found',
                    'status': 'failed'
                })
            except Exception as e:
                failed_sales.append({
                    'product_id': sale.get('product_id'),
                    'error': str(e),
                    'status': 'failed'
                })
        
        return Response({
            'processed_sales': processed_sales,
            'failed_sales': failed_sales,
            'total_processed': len(sales_data),
            'success_count': len(processed_sales),
            'failure_count': len(failed_sales)
        })
    
    @action(detail=False, methods=['get'])
    def store_locations(self, request):
        """Get all store locations for POS configuration"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        locations = []
        
        # Main supermarket
        locations.append({
            'id': supermarket.id,
            'type': 'supermarket',
            'name': supermarket.name,
            'address': supermarket.address,
            'phone': supermarket.phone,
            'is_active': supermarket.is_active
        })
        
        # Substores
        for substore in supermarket.substores.filter(is_active=True):
            locations.append({
                'id': substore.id,
                'type': 'substore',
                'name': substore.name,
                'address': substore.address,
                'phone': substore.phone,
                'parent_supermarket': supermarket.name,
                'is_active': substore.is_active
            })
        
        return Response({
            'locations': locations,
            'total_count': len(locations)
        })


class POSIntegrationViewSet(viewsets.ViewSet):
    """ViewSet for POS Integration operations"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'])
    def products_sync(self, request):
        """Sync products with POS system"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Get all products for this supermarket
        products = Product.objects.filter(
            Q(supermarket=supermarket) | Q(substore__supermarket=supermarket),
            is_active=True
        )
        
        # Simulate sync process
        synced_products = []
        for product in products:
            synced_products.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'barcode': product.barcode,
                'price': float(product.price),
                'current_stock': product.current_stock,
                'sync_status': 'success'
            })
        
        return Response({
            'message': 'Products synced successfully',
            'total_count': len(synced_products),
            'synced_products': synced_products,
            'sync_timestamp': timezone.now().isoformat()
        })
    
    @action(detail=False, methods=['post'])
    def stock_update_from_pos(self, request):
        """Update stock levels from POS system"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        updates = request.data.get('updates', [])
        if not updates:
            return Response({'error': 'No updates provided'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        successful_updates = []
        failed_updates = []
        
        for update in updates:
            try:
                product_id = update.get('product_id')
                new_stock = update.get('new_stock')
                transaction_type = update.get('transaction_type', 'ADJUSTMENT')
                reason = update.get('reason', 'POS System Update')
                
                # Find product
                product = Product.objects.get(
                    Q(id=product_id) & 
                    (Q(supermarket=supermarket) | Q(substore__supermarket=supermarket)) &
                    Q(is_active=True)
                )
                
                previous_stock = product.current_stock
                product.current_stock = new_stock
                product.save()
                
                # Create stock transaction record
                StockTransaction.objects.create(
                    product=product,
                    transaction_type=transaction_type,
                    quantity=abs(new_stock - previous_stock),
                    previous_stock=previous_stock,
                    new_stock=new_stock,
                    reason=reason,
                    user=request.user
                )
                
                successful_updates.append({
                    'product_id': product_id,
                    'previous_stock': previous_stock,
                    'new_stock': new_stock,
                    'status': 'success'
                })
                
            except Product.DoesNotExist:
                failed_updates.append({
                    'product_id': update.get('product_id'),
                    'error': 'Product not found',
                    'status': 'failed'
                })
            except Exception as e:
                failed_updates.append({
                    'product_id': update.get('product_id'),
                    'error': str(e),
                    'status': 'failed'
                })
        
        return Response({
            'successful_updates': successful_updates,
            'failed_updates': failed_updates,
            'total_processed': len(updates),
            'success_count': len(successful_updates),
            'failure_count': len(failed_updates)
        })
    
    @action(detail=False, methods=['post'])
    def sales_data_from_pos(self, request):
        """Receive sales data from POS system"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        sales_data = request.data.get('sales', [])
        if not sales_data:
            return Response({'error': 'No sales data provided'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        processed_sales = []
        failed_sales = []
        
        for sale in sales_data:
            try:
                product_id = sale.get('product_id')
                quantity_sold = sale.get('quantity_sold', 0)
                sale_price = sale.get('sale_price', 0)
                sale_timestamp = sale.get('timestamp')
                
                # Find product and update stock
                product = Product.objects.get(
                    Q(id=product_id) & 
                    (Q(supermarket=supermarket) | Q(substore__supermarket=supermarket)) &
                    Q(is_active=True)
                )
                
                if product.current_stock >= quantity_sold:
                    previous_stock = product.current_stock
                    product.current_stock -= quantity_sold
                    product.save()
                    
                    # Create stock transaction record
                    StockTransaction.objects.create(
                        product=product,
                        transaction_type='OUT',
                        quantity=quantity_sold,
                        previous_stock=previous_stock,
                        new_stock=product.current_stock,
                        reason=f'POS Sale - ${sale_price}',
                        user=request.user
                    )
                    
                    processed_sales.append({
                        'product_id': product_id,
                        'quantity_sold': quantity_sold,
                        'remaining_stock': product.current_stock,
                        'status': 'success'
                    })
                else:
                    failed_sales.append({
                        'product_id': product_id,
                        'error': f'Insufficient stock. Available: {product.current_stock}, Requested: {quantity_sold}',
                        'status': 'failed'
                    })
                    
            except Product.DoesNotExist:
                failed_sales.append({
                    'product_id': sale.get('product_id'),
                    'error': 'Product not found',
                    'status': 'failed'
                })
            except Exception as e:
                failed_sales.append({
                    'product_id': sale.get('product_id'),
                    'error': str(e),
                    'status': 'failed'
                })
        
        return Response({
            'processed_sales': processed_sales,
            'failed_sales': failed_sales,
            'total_processed': len(sales_data),
            'success_count': len(processed_sales),
            'failure_count': len(failed_sales)
        })
    
    @action(detail=False, methods=['get'])
    def store_locations(self, request):
        """Get all store locations for POS configuration"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        locations = []
        
        # Main supermarket
        locations.append({
            'id': supermarket.id,
            'type': 'supermarket',
            'name': supermarket.name,
            'address': supermarket.address,
            'phone': supermarket.phone,
            'is_active': supermarket.is_active
        })
        
        # Substores
        for substore in supermarket.substores.filter(is_active=True):
            locations.append({
                'id': substore.id,
                'type': 'substore',
                'name': substore.name,
                'address': substore.address,
                'phone': substore.phone,
                'parent_supermarket': supermarket.name,
                'is_active': substore.is_active
            })
        
        return Response({
            'locations': locations,
            'total_count': len(locations)
        })


class DashboardViewSet(viewsets.ViewSet):
    """ViewSet for Dashboard statistics and data"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get overall dashboard statistics"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Get all products for this supermarket
        products = Product.objects.filter(
            Q(supermarket=supermarket) | Q(substore__supermarket=supermarket),
            is_active=True
        )
        
        # Calculate statistics
        total_products = products.count()
        total_stock_value = sum(p.price * p.current_stock for p in products)
        low_stock_products = products.filter(current_stock__lte=F('minimum_stock')).count()
        
        # Expiry alerts
        today = timezone.now().date()
        expiring_soon = products.filter(
            expiry_date__lte=today + timedelta(days=7),
            expiry_date__gt=today
        ).count()
        
        expired_products = products.filter(expiry_date__lt=today).count()
        
        # Category breakdown
        category_stats = products.values('category').annotate(
            count=Count('id'),
            total_value=Sum(F('price') * F('current_stock'))
        ).order_by('-count')
        
        # Store breakdown
        store_stats = []
        
        # Main supermarket stats
        main_products = products.filter(supermarket=supermarket)
        store_stats.append({
            'type': 'supermarket',
            'name': supermarket.name,
            'product_count': main_products.count(),
            'total_value': sum(p.price * p.current_stock for p in main_products),
            'low_stock_count': main_products.filter(current_stock__lte=F('minimum_stock')).count()
        })
        
        # Substore stats
        for substore in supermarket.substores.filter(is_active=True):
            substore_products = products.filter(substore=substore)
            store_stats.append({
                'type': 'substore',
                'name': substore.name,
                'product_count': substore_products.count(),
                'total_value': sum(p.price * p.current_stock for p in substore_products),
                'low_stock_count': substore_products.filter(current_stock__lte=F('minimum_stock')).count()
            })
        
        return Response({
            'total_products': total_products,
            'total_stock_value': float(total_stock_value),
            'low_stock_products': low_stock_products,
            'expiring_soon': expiring_soon,
            'expired_products': expired_products,
            'category_breakdown': list(category_stats),
            'store_breakdown': store_stats,
            'total_stores': 1 + supermarket.substores.filter(is_active=True).count()
        })
    
    @action(detail=False, methods=['get'])
    def store_specific_stats(self, request):
        """Get statistics for a specific store"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        store_type = request.query_params.get('store_type')
        store_id = request.query_params.get('store_id')
        
        if not store_type or not store_id:
            return Response({'error': 'store_type and store_id are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if store_type == 'supermarket':
            products = Product.objects.filter(supermarket=supermarket, is_active=True)
            store_name = supermarket.name
        elif store_type == 'substore':
            try:
                substore = Substore.objects.get(id=store_id, supermarket=supermarket)
                products = Product.objects.filter(substore=substore, is_active=True)
                store_name = substore.name
            except Substore.DoesNotExist:
                return Response({'error': 'Substore not found'}, 
                              status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'error': 'Invalid store_type'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Calculate statistics for the specific store
        total_products = products.count()
        total_stock_value = sum(p.price * p.current_stock for p in products)
        low_stock_products = products.filter(current_stock__lte=F('minimum_stock')).count()
        
        # Expiry alerts
        today = timezone.now().date()
        expiring_soon = products.filter(
            expiry_date__lte=today + timedelta(days=7),
            expiry_date__gt=today
        ).count()
        
        expired_products = products.filter(expiry_date__lt=today).count()
        
        # Category breakdown
        category_stats = products.values('category').annotate(
            count=Count('id'),
            total_value=Sum(F('price') * F('current_stock'))
        ).order_by('-count')
        
        return Response({
            'store_name': store_name,
            'store_type': store_type,
            'total_products': total_products,
            'total_stock_value': float(total_stock_value),
            'low_stock_products': low_stock_products,
            'expiring_soon': expiring_soon,
            'expired_products': expired_products,
            'category_breakdown': list(category_stats)
        })
    
    @action(detail=False, methods=['get'])
    def alerts_summary(self, request):
        """Get summary of all alerts"""
        try:
            supermarket = Supermarket.objects.get(user=request.user)
        except Supermarket.DoesNotExist:
            return Response({'error': 'You must be registered as a supermarket'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Get all products for this supermarket
        products = Product.objects.filter(
            Q(supermarket=supermarket) | Q(substore__supermarket=supermarket),
            is_active=True
        )
        
        today = timezone.now().date()
        
        # Low stock alerts
        low_stock_products = products.filter(current_stock__lte=F('minimum_stock'))
        
        # Expiry alerts
        expiring_soon = products.filter(
            expiry_date__lte=today + timedelta(days=7),
            expiry_date__gt=today
        )
        
        expired_products = products.filter(expiry_date__lt=today)
        
        # Out of stock
        out_of_stock = products.filter(current_stock=0)
        
        return Response({
            'low_stock': {
                'count': low_stock_products.count(),
                'products': [
                    {
                        'id': p.id,
                        'name': p.name,
                        'current_stock': p.current_stock,
                        'minimum_stock': p.minimum_stock,
                        'store_name': p.substore.name if p.substore else supermarket.name
                    } for p in low_stock_products[:10]  # Limit to 10 for performance
                ]
            },
            'expiring_soon': {
                'count': expiring_soon.count(),
                'products': [
                    {
                        'id': p.id,
                        'name': p.name,
                        'expiry_date': p.expiry_date,
                        'days_until_expiry': (p.expiry_date - today).days,
                        'store_name': p.substore.name if p.substore else supermarket.name
                    } for p in expiring_soon[:10]
                ]
            },
            'expired': {
                'count': expired_products.count(),
                'products': [
                    {
                        'id': p.id,
                        'name': p.name,
                        'expiry_date': p.expiry_date,
                        'days_expired': (today - p.expiry_date).days,
                        'store_name': p.substore.name if p.substore else supermarket.name
                    } for p in expired_products[:10]
                ]
            },
            'out_of_stock': {
                'count': out_of_stock.count(),
                'products': [
                    {
                        'id': p.id,
                        'name': p.name,
                        'store_name': p.substore.name if p.substore else supermarket.name
                    } for p in out_of_stock[:10]
                ]
            }
        })